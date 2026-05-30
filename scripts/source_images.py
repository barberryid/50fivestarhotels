import csv
import json
import re
import time
from datetime import date
from pathlib import Path
from urllib.parse import urlencode
from urllib.request import Request, urlopen
from urllib.error import URLError, HTTPError

from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parents[1]
TODAY = date.today().isoformat()
USER_AGENT = "50fivestarhotels-image-audit/1.0 (royalty-free travel image research)"

HOTELS = [
    (1, "Hotel Majapahit Surabaya - MGallery Collection", "hotel-majapahit-surabaya", "Surabaya", "Indonesia", [
        ("Hero", "Hotel Majapahit Surabaya heritage building exterior", "heritage-building"),
        ("Hotel", "Surabaya colonial architecture heritage hotel garden", "colonial-architecture"),
        ("City", "Surabaya old town heritage architecture", "surabaya-old-town"),
        ("Attraction", "House of Sampoerna Surabaya", "house-of-sampoerna"),
        ("Food", "East Java food market Surabaya", "east-java-food"),
        ("Secondary", "Heroes Monument Surabaya", "heroes-monument"),
    ]),
    (2, "The Phoenix Hotel Yogyakarta - Handwritten Collection", "phoenix-hotel-yogyakarta", "Yogyakarta", "Indonesia", [
        ("Hero", "The Phoenix Hotel Yogyakarta", "phoenix-hotel"),
        ("Hotel", "Yogyakarta heritage architecture", "heritage-architecture"),
        ("City", "Malioboro Yogyakarta street", "malioboro"),
        ("Attraction", "Prambanan temple Yogyakarta", "prambanan"),
        ("Food", "Yogyakarta gudeg market", "gudeg-market"),
        ("Secondary", "Borobudur temple Java", "borobudur"),
    ]),
    (3, "The Majestic Hotel Kuala Lumpur, Autograph Collection", "majestic-hotel-kuala-lumpur", "Kuala Lumpur", "Malaysia", [
        ("Hero", "Majestic Hotel Kuala Lumpur", "majestic-hotel"),
        ("Hotel", "Kuala Lumpur colonial architecture", "colonial-architecture"),
        ("City", "Kuala Lumpur skyline", "kl-skyline"),
        ("Attraction", "Sultan Abdul Samad Building Kuala Lumpur", "sultan-abdul-samad"),
        ("Food", "Kuala Lumpur food market", "kl-food-market"),
        ("Secondary", "Merdeka Square Kuala Lumpur", "merdeka-square"),
    ]),
    (4, "The Raweekanlaya Bangkok Wellness Cuisine Resort", "raweekanlaya-bangkok", "Bangkok", "Thailand", [
        ("Hero", "The Raweekanlaya Bangkok", "raweekanlaya"),
        ("Hotel", "old Bangkok heritage house garden", "old-bangkok-house"),
        ("City", "Chao Phraya River Bangkok", "chao-phraya"),
        ("Attraction", "Wat Pho Bangkok", "wat-pho"),
        ("Food", "Bangkok street food market", "street-food"),
        ("Secondary", "Grand Palace Bangkok", "grand-palace"),
    ]),
    (5, "Stamba Hotel", "stamba-hotel-tbilisi", "Tbilisi", "Georgia", [
        ("Hero", "Stamba Hotel Tbilisi", "stamba-hotel"),
        ("Hotel", "Tbilisi design architecture", "design-architecture"),
        ("City", "Tbilisi old town architecture", "old-town"),
        ("Attraction", "Narikala Fortress Tbilisi", "narikala"),
        ("Food", "Georgian food khachapuri Tbilisi", "georgian-food"),
        ("Secondary", "Rustaveli Avenue Tbilisi", "rustaveli"),
    ]),
    (6, "Angkor Aurora", "angkor-aurora-siem-reap", "Siem Reap", "Cambodia", [
        ("Hero", "Angkor Aurora Siem Reap", "angkor-aurora"),
        ("Hotel", "Siem Reap boutique pool hotel", "boutique-pool"),
        ("City", "Siem Reap Cambodia street", "siem-reap"),
        ("Attraction", "Angkor Wat Cambodia", "angkor-wat"),
        ("Food", "Cambodian food market Siem Reap", "cambodian-food"),
        ("Secondary", "Ta Prohm Cambodia", "ta-prohm"),
    ]),
    (7, "The Royal Surakarta Heritage - Handwritten Collection", "royal-surakarta-heritage", "Solo / Surakarta", "Indonesia", [
        ("Hero", "Royal Surakarta Heritage Solo", "royal-surakarta"),
        ("Hotel", "Surakarta Javanese palace interior", "javanese-palace"),
        ("City", "Surakarta Solo city Indonesia", "surakarta-city"),
        ("Attraction", "Keraton Surakarta", "keraton-surakarta"),
        ("Food", "Solo Indonesia food market", "solo-food"),
        ("Secondary", "Pura Mangkunegaran Surakarta", "mangkunegaran"),
    ]),
    (8, "Pleiada Boutique Hotel & Spa", "pleiada-boutique-hotel-iasi", "Iasi", "Romania", [
        ("Hero", "Pleiada Boutique Hotel Iasi", "pleiada-hotel"),
        ("Hotel", "Iasi Romania boutique architecture", "iasi-architecture"),
        ("City", "Iasi Romania city", "iasi-city"),
        ("Attraction", "Palace of Culture Iasi", "palace-of-culture"),
        ("Food", "Romanian food Iasi", "romanian-food"),
        ("Secondary", "Copou Park Iasi", "copou-park"),
    ]),
    (9, "Steigenberger Resort Achti Luxor", "steigenberger-resort-achti-luxor", "Luxor", "Egypt", [
        ("Hero", "Steigenberger Resort Achti Luxor Nile", "achti-luxor"),
        ("Hotel", "Luxor Nile resort garden", "nile-resort"),
        ("City", "Luxor Nile Egypt", "luxor-nile"),
        ("Attraction", "Karnak Temple Luxor", "karnak-temple"),
        ("Food", "Egyptian food Luxor market", "egyptian-food"),
        ("Secondary", "Valley of the Kings Luxor", "valley-of-the-kings"),
    ]),
    (10, "Silk Path Grand Hue Hotel & Spa", "silk-path-grand-hue", "Hue", "Vietnam", [
        ("Hero", "Silk Path Grand Hue Hotel", "silk-path-grand"),
        ("Hotel", "Hue Vietnam colonial hotel architecture", "hue-architecture"),
        ("City", "Perfume River Hue Vietnam", "perfume-river"),
        ("Attraction", "Imperial City Hue Vietnam", "imperial-city"),
        ("Food", "Hue Vietnamese food market", "hue-food"),
        ("Secondary", "Thien Mu Pagoda Hue", "thien-mu"),
    ]),
    (11, "The Hermitage, A Tribute Portfolio Hotel, Jakarta", "hermitage-jakarta", "Jakarta", "Indonesia", [
        ("Hero", "The Hermitage Jakarta Menteng", "hermitage-jakarta"),
        ("Hotel", "Menteng Jakarta heritage architecture", "menteng-heritage"),
        ("City", "Jakarta skyline", "jakarta-skyline"),
        ("Attraction", "National Monument Jakarta", "monas"),
        ("Food", "Jakarta Indonesian food market satay nasi goreng", "jakarta-food"),
        ("Secondary", "Kota Tua Jakarta", "kota-tua"),
    ]),
    (12, "Hyatt Regency Tashkent", "hyatt-regency-tashkent", "Tashkent", "Uzbekistan", [
        ("Hero", "Hyatt Regency Tashkent", "hyatt-regency"),
        ("Hotel", "Tashkent modern architecture", "modern-architecture"),
        ("City", "Tashkent city Uzbekistan", "tashkent-city"),
        ("Attraction", "Hazrati Imam Complex Tashkent", "hazrati-imam"),
        ("Food", "Uzbek food market Tashkent", "uzbek-food"),
        ("Secondary", "Tashkent metro", "tashkent-metro"),
    ]),
    (13, "Grand Hotel Yerevan", "grand-hotel-yerevan", "Yerevan", "Armenia", [
        ("Hero", "Grand Hotel Yerevan", "grand-hotel"),
        ("Hotel", "Yerevan Republic Square architecture", "republic-square"),
        ("City", "Yerevan Mount Ararat", "mount-ararat"),
        ("Attraction", "Cascade Complex Yerevan", "cascade"),
        ("Food", "Armenian food Yerevan cafe", "armenian-food"),
        ("Secondary", "Republic Square Yerevan", "republic-square-night"),
    ]),
    (14, "Sofitel Marrakech Palais Imperial & Spa", "sofitel-marrakech", "Marrakech", "Morocco", [
        ("Hero", "Sofitel Marrakech Palais Imperial", "sofitel"),
        ("Hotel", "Marrakech Hivernage luxury garden", "hivernage-garden"),
        ("City", "Marrakech medina Morocco", "medina"),
        ("Attraction", "Koutoubia Mosque Marrakech", "koutoubia"),
        ("Food", "Marrakech souk food", "souk-food"),
        ("Secondary", "Jardin Majorelle Marrakech", "jardin-majorelle"),
    ]),
    (15, "Boton Blue Hotel & Spa", "boton-blue-nha-trang", "Nha Trang", "Vietnam", [
        ("Hero", "Boton Blue Hotel Nha Trang", "boton-blue"),
        ("Hotel", "Nha Trang sea view hotel", "sea-view"),
        ("City", "Nha Trang beach Vietnam", "nha-trang-beach"),
        ("Attraction", "Po Nagar Cham Towers Nha Trang", "po-nagar"),
        ("Food", "Nha Trang seafood market", "seafood"),
        ("Secondary", "Hon Chong Nha Trang", "hon-chong"),
    ]),
    (16, "I'M Hotel Makati", "im-hotel-makati", "Makati / Manila", "Philippines", [
        ("Hero", "I'M Hotel Makati", "im-hotel"),
        ("Hotel", "Makati rooftop pool skyline", "rooftop-skyline"),
        ("City", "Makati skyline Manila", "makati-skyline"),
        ("Attraction", "Ayala Triangle Makati", "ayala-triangle"),
        ("Food", "Poblacion Makati food nightlife", "poblacion-food"),
        ("Secondary", "Manila Bay Philippines", "manila-bay"),
    ]),
    (17, "Malak Regency Hotel", "malak-regency-sarajevo", "Sarajevo", "Bosnia and Herzegovina", [
        ("Hero", "Malak Regency Hotel Sarajevo", "malak-regency"),
        ("Hotel", "Sarajevo luxury hotel mountains", "sarajevo-mountains"),
        ("City", "Sarajevo old town Bascarsija", "bascarsija"),
        ("Attraction", "Latin Bridge Sarajevo", "latin-bridge"),
        ("Food", "Sarajevo cevapi Bascarsija", "cevapi"),
        ("Secondary", "Sarajevo mountains", "mountains"),
    ]),
    (18, "Real Marina Hotel & Spa", "real-marina-olhao", "Olhao", "Portugal", [
        ("Hero", "Real Marina Hotel Olhao", "real-marina"),
        ("Hotel", "Olhao marina Algarve", "olhao-marina"),
        ("City", "Olhao Portugal old town", "olhao-old-town"),
        ("Attraction", "Ria Formosa Olhao", "ria-formosa"),
        ("Food", "Olhao seafood market", "seafood-market"),
        ("Secondary", "Culatra Island Algarve", "culatra-island"),
    ]),
    (19, "Graffit Gallery Design Hotel", "graffit-gallery-varna", "Varna", "Bulgaria", [
        ("Hero", "Graffit Gallery Design Hotel Varna", "graffit-gallery"),
        ("Hotel", "Varna design hotel architecture", "design-architecture"),
        ("City", "Varna Bulgaria Black Sea", "varna-black-sea"),
        ("Attraction", "Roman Baths Varna", "roman-baths"),
        ("Food", "Varna Bulgaria food market", "varna-food"),
        ("Secondary", "Varna Sea Garden", "sea-garden"),
    ]),
    (20, "Millennium Downtown Abu Dhabi", "millennium-downtown-abu-dhabi", "Abu Dhabi", "UAE", [
        ("Hero", "Millennium Downtown Abu Dhabi", "millennium-downtown"),
        ("Hotel", "Abu Dhabi skyline downtown", "abu-dhabi-skyline"),
        ("City", "Abu Dhabi Corniche", "corniche"),
        ("Attraction", "Sheikh Zayed Grand Mosque Abu Dhabi", "grand-mosque"),
        ("Food", "Abu Dhabi Emirati food market", "emirati-food"),
        ("Secondary", "Qasr Al Watan Abu Dhabi", "qasr-al-watan"),
    ]),
]

BAD_SOURCE_PATTERNS = ["booking.com", "agoda", "expedia", "tripadvisor", "kayak", "trivago", "instagram", "pinterest"]
LICENCE_NOTES = [
    ["Wikimedia Commons", "Creative Commons / Public Domain, image-specific", "https://commons.wikimedia.org/wiki/Commons:Licensing", "Yes, where licence permits", "Often yes for Creative Commons; no for public domain", "Respect image-specific licence and attribution terms; avoid non-commercial or unclear licences.", "All downloaded files in this audit came from Commons imageinfo metadata."],
]

used_pages = set()
rejected = []
rows = []


def slugify(text):
    text = text.lower().replace("'", "")
    text = re.sub(r"[^a-z0-9]+", "-", text).strip("-")
    return text or "image"


def api_json(params):
    params = {"format": "json", "formatversion": "2", **params}
    url = "https://commons.wikimedia.org/w/api.php?" + urlencode(params)
    req = Request(url, headers={"User-Agent": USER_AGENT})
    for attempt in range(5):
        try:
            with urlopen(req, timeout=40) as response:
                return json.loads(response.read().decode("utf-8"))
        except HTTPError as exc:
            if exc.code == 429 and attempt < 4:
                time.sleep(12 + attempt * 8)
                continue
            raise


def search_commons(query, limit=18):
    data = api_json({
        "action": "query",
        "generator": "search",
        "gsrsearch": query,
        "gsrnamespace": 6,
        "gsrlimit": limit,
        "prop": "imageinfo",
        "iiprop": "url|size|mime|extmetadata|commonmetadata|thumburl",
        "iiurlwidth": 2200,
    })
    return data.get("query", {}).get("pages", [])


def clean_meta_value(meta, key):
    val = meta.get(key, {}).get("value") if isinstance(meta, dict) else None
    if not val:
        return "Unknown / not listed"
    val = re.sub(r"<[^>]+>", "", str(val))
    return " ".join(val.split()) or "Unknown / not listed"


def license_ok(short_name, license_url):
    text = f"{short_name} {license_url}".lower()
    if any(x in text for x in ["noncommercial", "non-commercial", "-nc", "by-nc", "no commercial"]):
        return False
    if any(x in text for x in ["cc by", "cc0", "public domain", "pd-old", "attribution", "share alike", "by-sa", "gfdl", "free art license", "fal"]):
        return True
    return False


def score_page(page):
    info = (page.get("imageinfo") or [{}])[0]
    title = page.get("title", "")
    mime = info.get("mime", "")
    width = int(info.get("width") or 0)
    height = int(info.get("height") or 0)
    ext = info.get("extmetadata") or {}
    licence = clean_meta_value(ext, "LicenseShortName")
    license_url = clean_meta_value(ext, "LicenseUrl")
    source = clean_meta_value(ext, "ImageDescription") + " " + clean_meta_value(ext, "Credit")
    lower_blob = f"{title} {source} {info.get('descriptionurl','')}".lower()
    if mime not in ("image/jpeg", "image/png", "image/webp"):
        return None, "Unsupported mime type"
    if width < 1000 or height < 650:
        return None, "Too low resolution"
    if any(p in lower_blob for p in BAD_SOURCE_PATTERNS):
        return None, "Banned source reference in metadata"
    if not license_ok(licence, license_url):
        return None, f"Licence not clearly commercial-safe: {licence}"
    score = 0
    if width >= height:
        score += 25
    if width >= 1600:
        score += 20
    if "watermark" in lower_blob:
        score -= 50
    if "map" in lower_blob or "logo" in lower_blob or "seal" in lower_blob:
        score -= 20
    if "quality image" in lower_blob or "featured picture" in lower_blob:
        score += 10
    score += min(width / 400, 10)
    return score, None


def choose_image(query, broader_queries):
    candidates = []
    for q in [query] + broader_queries:
        try:
            pages = search_commons(q)
        except Exception as exc:
            rejected.append([query, q, "Wikimedia Commons", "", "Search failed", "No", "No", str(exc)])
            continue
        for page in pages:
            info = (page.get("imageinfo") or [{}])[0]
            page_url = info.get("descriptionurl", "")
            if not page_url or page_url in used_pages:
                continue
            score, reason = score_page(page)
            info_blob = (
                page.get("title", "")
                + " "
                + clean_meta_value(info.get("extmetadata") or {}, "ImageDescription")
                + " "
                + clean_meta_value(info.get("extmetadata") or {}, "Credit")
            ).lower()
            stop_words = {
                "hotel", "resort", "boutique", "travel", "city", "food", "market",
                "street", "image", "architecture", "heritage", "garden", "luxury",
                "indonesia", "vietnam", "romania", "egypt", "morocco",
                "philippines", "portugal", "bulgaria", "malaysia", "thailand",
                "georgia", "cambodia", "uzbekistan", "armenia", "bosnia",
                "emirates", "united", "arab",
            }
            tokens = [tok for tok in re.findall(r"[a-z]{4,}", q.lower()) if tok not in stop_words]
            if score is not None and tokens and not any(tok in info_blob for tok in tokens[:5]):
                score, reason = None, "Wrong location or weak relevance"
            if score is None:
                rejected.append([query, page.get("title", q), "Wikimedia Commons", page_url, reason, "Yes" if "Licence" in reason else "No", "Yes" if "resolution" in reason or "mime" in reason else "No", "Rejected by automated QC"])
                continue
            candidates.append((score, page, q))
        if candidates:
            break
        time.sleep(0.25)
    if not candidates:
        return None, None
    candidates.sort(key=lambda item: item[0], reverse=True)
    page = candidates[0][1]
    used_pages.add((page.get("imageinfo") or [{}])[0].get("descriptionurl", ""))
    return page, candidates[0][2]


def download(url, dest):
    req = Request(url, headers={"User-Agent": USER_AGENT})
    for attempt in range(5):
        try:
            with urlopen(req, timeout=90) as response:
                dest.write_bytes(response.read())
                return
        except HTTPError as exc:
            if exc.code == 429 and attempt < 4:
                time.sleep(20 + attempt * 10)
                continue
            raise


def optimize(src, dest):
    with Image.open(src) as im:
        im = im.convert("RGB")
        if im.width > 2200:
            ratio = 2200 / im.width
            im = im.resize((2200, int(im.height * ratio)), Image.Resampling.LANCZOS)
        quality = 82
        while True:
            im.save(dest, "WEBP", quality=quality, method=6)
            if dest.stat().st_size <= 1_500_000 or quality <= 68:
                break
            quality -= 4
        return im.width, im.height, dest.stat().st_size


def ext_from_mime(mime, url):
    if mime == "image/png":
        return "png"
    if mime == "image/webp":
        return "webp"
    return "jpg"


def attribution_text(title, creator, licence, page_url):
    if "public domain" in licence.lower() or licence.upper().startswith("CC0"):
        return f"Image: {title} by {creator}, public domain via Wikimedia Commons."
    return f"Photo: {title} by {creator}, licensed under {licence} via Wikimedia Commons. Cropped/resized only if needed and converted to WebP."


def build_outputs():
    for rank, hotel, slug, city, country, slots in HOTELS:
        hotel_dir = ROOT / "public" / "images" / "hotels" / slug
        original_dir = hotel_dir / "original"
        web_dir = hotel_dir / "web"
        original_dir.mkdir(parents=True, exist_ok=True)
        web_dir.mkdir(parents=True, exist_ok=True)
        for idx, (image_type, subject, short_subject) in enumerate(slots, 1):
            broader = [
                subject.replace(hotel, city),
                f"{city} {country} {short_subject.replace('-', ' ')}",
                f"{city} {country} travel",
                f"{country} {short_subject.replace('-', ' ')}",
            ]
            page, matched_query = choose_image(subject, broader)
            image_id = f"{slug}-{idx:02d}"
            if not page:
                rows.append({
                    "ID": image_id, "Hotel Rank": rank, "Hotel Name": hotel, "Hotel Slug": slug, "City": city, "Country": country,
                    "Image Slot": idx, "Image Type": image_type, "Subject": subject, "File Name Original": "", "File Name Web": "",
                    "Local Path Original": "", "Local Path Web": "", "Source Site": "Wikimedia Commons", "Source Page URL": "", "Direct Download URL": "",
                    "Photographer / Creator": "Unknown / not listed", "Photographer Profile URL": "", "Licence": "", "Licence URL": "", "Attribution Required?": "",
                    "Recommended Attribution Text": "", "Commercial Use Allowed?": "Unclear", "Modification Allowed?": "Unclear", "Verification Date": TODAY,
                    "Notes": "No safe image found by automated Commons search; manual replacement recommended. Hotel-specific image not found under safe licence", "Safe to Use?": "Review Needed",
                    "Width": "", "Height": "", "Web File Size Bytes": ""
                })
                continue
            info = page["imageinfo"][0]
            meta = info.get("extmetadata") or {}
            title = page.get("title", "").replace("File:", "")
            creator = clean_meta_value(meta, "Artist")
            licence = clean_meta_value(meta, "LicenseShortName")
            licence_url = clean_meta_value(meta, "LicenseUrl")
            page_url = info.get("descriptionurl", "")
            download_url = info.get("thumburl") or info.get("url", "")
            ext = ext_from_mime(info.get("mime", ""), download_url)
            base = f"{slug}-{idx:02d}-{image_type.lower()}-{slugify(short_subject)}"
            original_name = f"{base}.{ext}"
            web_name = f"{base}.webp"
            original_path = original_dir / original_name
            web_path = web_dir / web_name
            try:
                if not original_path.exists():
                    download(download_url, original_path)
                width, height, web_size = optimize(original_path, web_path)
                safe = "Yes"
                commercial = "Yes"
                modification = "Yes"
            except Exception as exc:
                width = height = web_size = ""
                safe = "Review Needed"
                commercial = modification = "Unclear"
                rejected.append([hotel, subject, "Wikimedia Commons", page_url, "Download or conversion failed", "No", "Yes", str(exc)])
            notes = f"Matched Commons query: {matched_query}."
            if image_type in ("Hero", "Hotel") and hotel.lower().split()[0] not in title.lower():
                notes += " Hotel-specific image not found under safe licence; using destination/neighbourhood alternative."
            rows.append({
                "ID": image_id, "Hotel Rank": rank, "Hotel Name": hotel, "Hotel Slug": slug, "City": city, "Country": country,
                "Image Slot": idx, "Image Type": image_type, "Subject": title or subject, "File Name Original": original_name, "File Name Web": web_name,
                "Local Path Original": str(original_path.relative_to(ROOT)).replace("\\", "/"), "Local Path Web": "/" + str(web_path.relative_to(ROOT / "public")).replace("\\", "/"),
                "Source Site": "Wikimedia Commons", "Source Page URL": page_url, "Direct Download URL": download_url,
                "Photographer / Creator": creator, "Photographer Profile URL": "", "Licence": licence, "Licence URL": licence_url, "Attribution Required?": "No" if ("public domain" in licence.lower() or licence.upper().startswith("CC0")) else "Yes",
                "Recommended Attribution Text": attribution_text(title, creator, licence, page_url), "Commercial Use Allowed?": commercial, "Modification Allowed?": modification,
                "Verification Date": TODAY, "Notes": notes, "Safe to Use?": safe, "Width": width, "Height": height, "Web File Size Bytes": web_size
            })
            print((f"{image_id}: {safe} - {title[:90]}").encode("ascii", "replace").decode("ascii"))
            time.sleep(0.3)


def write_csv_json():
    data_dir = ROOT / "src" / "data"
    data_dir.mkdir(parents=True, exist_ok=True)
    headers = list(rows[0].keys())
    csv_path = data_dir / "imageAttributions.csv"
    with csv_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)
    def camel(row):
        mapping = {
            "ID": "id", "Hotel Rank": "hotelRank", "Hotel Name": "hotelName", "Hotel Slug": "hotelSlug", "City": "city", "Country": "country",
            "Image Slot": "imageSlot", "Image Type": "imageType", "Subject": "subject", "File Name Original": "fileNameOriginal", "File Name Web": "fileNameWeb",
            "Local Path Original": "localPathOriginal", "Local Path Web": "localPathWeb", "Source Site": "sourceSite", "Source Page URL": "sourcePageUrl",
            "Direct Download URL": "directDownloadUrl", "Photographer / Creator": "photographer", "Photographer Profile URL": "photographerProfileUrl",
            "Licence": "license", "Licence URL": "licenseUrl", "Attribution Required?": "attributionRequired", "Recommended Attribution Text": "recommendedAttributionText",
            "Commercial Use Allowed?": "commercialUseAllowed", "Modification Allowed?": "modificationAllowed", "Verification Date": "verificationDate", "Notes": "notes", "Safe to Use?": "safeToUse",
            "Width": "width", "Height": "height", "Web File Size Bytes": "webFileSizeBytes"
        }
        out = {}
        for k, v in row.items():
            key = mapping[k]
            if k == "Attribution Required?":
                out[key] = True if v == "Yes" else False if v == "No" else None
            else:
                out[key] = v
        return out
    (data_dir / "imageAttributions.json").write_text(json.dumps([camel(r) for r in rows], ensure_ascii=False, indent=2), encoding="utf-8")


def summary_rows():
    summary = []
    for rank, hotel, slug, city, country, _ in HOTELS:
        hotel_rows = [r for r in rows if r["Hotel Slug"] == slug]
        safe_rows = [r for r in hotel_rows if r["Safe to Use?"] == "Yes"]
        missing_specific = [r for r in hotel_rows[:2] if "Hotel-specific image not found" in r["Notes"]]
        concerns = "; ".join(r["ID"] for r in hotel_rows if r["Safe to Use?"] != "Yes") or "None"
        missing = []
        for t in ["Hero", "Hotel", "City", "Attraction", "Food", "Secondary"]:
            if not any(r["Image Type"] == t and r["Safe to Use?"] == "Yes" for r in hotel_rows):
                missing.append(t)
        strength = "Excellent" if len(safe_rows) == 6 and not missing_specific else "Strong" if len(safe_rows) == 6 else "Acceptable" if len(safe_rows) >= 5 else "Needs Replacement"
        summary.append({
            "Hotel Rank": rank, "Hotel Name": hotel, "City": city, "Country": country, "Hotel Slug": slug,
            "Number of Images Found": len([r for r in hotel_rows if r["File Name Web"]]), "Number Safe to Use": len(safe_rows),
            "Has Hotel-Specific Image?": "No" if missing_specific else "Yes", "Has City Image?": "Yes" if any(r["Image Type"] == "City" and r["Safe to Use?"] == "Yes" for r in hotel_rows) else "No",
            "Has Major Attraction Image?": "Yes" if any(r["Image Type"] == "Attraction" and r["Safe to Use?"] == "Yes" for r in hotel_rows) else "No", "Has Food/Culture Image?": "Yes" if any(r["Image Type"] == "Food" and r["Safe to Use?"] == "Yes" for r in hotel_rows) else "No",
            "Best Hero Image ID": (next((r["ID"] for r in hotel_rows if r["Image Slot"] == 1 and r["Safe to Use?"] == "Yes"), "")),
            "Any Licensing Concerns": concerns, "Missing Items": ", ".join(missing) or "None", "Overall Visual Strength": strength,
            "Notes": "Hotel-specific hero/stay image needs manual review." if missing_specific else "Coverage complete with Commons-sourced images."
        })
    return summary


def write_workbook():
    out_path = ROOT / "research" / "image-attributions.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Image Attributions"
    attr_headers = list(rows[0].keys())
    ws.append(attr_headers)
    for row in rows:
        ws.append([row[h] for h in attr_headers])
    add_table_style(ws, "ImageAttributionsTable")

    ws2 = wb.create_sheet("Hotel Image Summary")
    srows = summary_rows()
    sheaders = list(srows[0].keys())
    ws2.append(sheaders)
    for row in srows:
        ws2.append([row[h] for h in sheaders])
    add_table_style(ws2, "HotelImageSummaryTable")

    ws3 = wb.create_sheet("Rejected Images")
    rej_headers = ["Hotel", "Candidate Subject", "Source Site", "Source URL", "Reason Rejected", "Licence Issue?", "Quality Issue?", "Notes"]
    ws3.append(rej_headers)
    for row in rejected:
        ws3.append(row)
    add_table_style(ws3, "RejectedImagesTable")

    ws4 = wb.create_sheet("Licence Notes")
    lic_headers = ["Source Site", "General Licence Name", "Licence URL", "Commercial Use Allowed", "Attribution Required", "Important Restrictions", "Notes"]
    ws4.append(lic_headers)
    for row in LICENCE_NOTES:
        ws4.append(row)
    add_table_style(ws4, "LicenceNotesTable")

    for sheet in wb.worksheets:
        sheet.freeze_panes = "A2"
        for col_idx, col in enumerate(sheet.columns, 1):
            max_len = max(len(str(cell.value or "")) for cell in col[:80])
            sheet.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 48)
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E5F")
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    wb.save(out_path)


def add_table_style(ws, name):
    max_row = max(ws.max_row, 1)
    max_col = max(ws.max_column, 1)
    ref = f"A1:{get_column_letter(max_col)}{max_row}"
    tab = Table(displayName=name, ref=ref)
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    ws.add_table(tab)


def write_audit():
    srows = summary_rows()
    downloaded = len([r for r in rows if r["File Name Web"]])
    safe = len([r for r in rows if r["Safe to Use?"] == "Yes"])
    review = len([r for r in rows if r["Safe to Use?"] == "Review Needed"])
    missing_hotel = [s["Hotel Name"] for s in srows if s["Has Hotel-Specific Image?"] == "No"]
    excellent = [s["Hotel Name"] for s in srows if s["Overall Visual Strength"] == "Excellent"]
    replace = [s["Hotel Name"] for s in srows if s["Overall Visual Strength"] in ("Weak", "Needs Replacement")]
    lines = [
        "# 50 Five-Star Hotels Image Audit", "", "## Summary", "",
        f"- Total hotels: 20", f"- Required images: 120", f"- Images downloaded: {downloaded}", f"- Images safe to use: {safe}", f"- Images needing review: {review}",
        f"- Hotels missing hotel-specific image: {len(missing_hotel)}" + (f" ({'; '.join(missing_hotel)})" if missing_hotel else ""),
        f"- Hotels with excellent visual coverage: {len(excellent)}", f"- Hotels needing replacement images: {len(replace)}" + (f" ({'; '.join(replace)})" if replace else ""),
        "", "## Notes by hotel", ""
    ]
    for s in srows:
        hotel_rows = [r for r in rows if r["Hotel Slug"] == s["Hotel Slug"]]
        lines += [
            f"### {s['Hotel Name']}", "",
            f"- Images found: {s['Number of Images Found']} / 6; safe to use: {s['Number Safe to Use']} / 6",
            f"- Strongest image: {s['Best Hero Image ID'] or 'Needs manual selection'}",
            f"- Licence concerns: {s['Any Licensing Concerns']}",
            f"- Missing items: {s['Missing Items']}",
            f"- Notes: {s['Notes']}",
            ""
        ]
        for r in hotel_rows:
            lines.append(f"  - {r['ID']} ({r['Image Type']}): {r['Safe to Use?']} - {r['Subject']}")
        lines.append("")
    (ROOT / "research" / "image-audit.md").write_text("\n".join(lines), encoding="utf-8")


if __name__ == "__main__":
    build_outputs()
    if not rows:
        raise SystemExit("No rows generated")
    write_csv_json()
    write_workbook()
    write_audit()
    print(f"DONE rows={len(rows)} safe={len([r for r in rows if r['Safe to Use?']=='Yes'])} review={len([r for r in rows if r['Safe to Use?']=='Review Needed'])}")



